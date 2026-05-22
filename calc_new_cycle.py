# -*- coding: utf-8 -*-
"""
新公式计算各部门回款周期
公式：回款周期 = (欠款加权天数 + 回款加权天数) / (欠款总额 + 认款协同金额)
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from datetime import datetime
from collections import defaultdict

REGION = '中西部大区'
EXCEL = r'C:\Users\wm881\Downloads\业绩 欠款看板.xlsx'

wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)

# ==== 欠款数据 ====
# Col2=欠款天数, Col25=欠款金额, Col32=一级部门, Col34=三级部门
ws_debt = wb['欠款数据 ']
dept_debt_weighted = defaultdict(lambda: {'weighted_days': 0.0, 'total_debt': 0.0})

for r in range(2, ws_debt.max_row + 1):
    region = ws_debt.cell(r, 32).value
    if region != REGION:
        continue
    dept = ws_debt.cell(r, 34).value
    days = ws_debt.cell(r, 2).value
    amt = ws_debt.cell(r, 25).value  # 欠款金额（可能为负）
    
    if not dept:
        continue
    try:
        days = float(days) if days else 0
        amt = float(amt) if amt else 0
    except:
        continue
    
    # 使用绝对值
    dept_debt_weighted[dept]['weighted_days'] += abs(amt) * days
    dept_debt_weighted[dept]['total_debt'] += abs(amt)

print("=== 欠款加权计算 ===")
for dept, v in dept_debt_weighted.items():
    print(f"  {dept}: 欠款总额={v['total_debt']/10000:.2f}万, 加权={v['weighted_days']/10000:.0f}万元天")

# ==== 认款数据 ====
# Col12=业绩日期, Col16=认款时间, Col20=认款协同金额, Col27=一级部门, Col29=三级部门
ws_rec = wb['26财年Q1认款数据']
dept_rec_weighted = defaultdict(lambda: {'weighted_days': 0.0, 'total_rec': 0.0})

for r in range(2, ws_rec.max_row + 1):
    region = ws_rec.cell(r, 27).value
    if region != REGION:
        continue
    dept = ws_rec.cell(r, 29).value
    perf_date = ws_rec.cell(r, 12).value   # 业绩日期
    rec_time = ws_rec.cell(r, 16).value    # 认款时间
    amt = ws_rec.cell(r, 20).value         # 认款协同金额

    if not dept:
        continue
    try:
        amt = float(amt) if amt else 0
    except:
        continue
    
    # 计算回款天数 = 认款时间 - 业绩日期
    if isinstance(perf_date, datetime) and isinstance(rec_time, datetime):
        rec_days = (rec_time - perf_date).days
        rec_days = max(0, rec_days)  # 不允许负天数
    else:
        rec_days = 0
    
    dept_rec_weighted[dept]['weighted_days'] += amt * rec_days
    dept_rec_weighted[dept]['total_rec'] += amt

print("\n=== 认款加权计算 ===")
for dept, v in dept_rec_weighted.items():
    print(f"  {dept}: 认款总额={v['total_rec']/10000:.2f}万, 回款加权={v['weighted_days']/10000:.0f}万元天")

# ==== 计算回款周期 ====
print("\n=== 各部门回款周期（新公式）===")
print("公式：(欠款加权天数 + 回款加权天数) / (欠款总额 + 认款协同金额)\n")

all_depts = sorted(set(list(dept_debt_weighted.keys()) + list(dept_rec_weighted.keys())))
total_debt_weighted = 0
total_rec_weighted = 0
total_debt_amt = 0
total_rec_amt = 0

results = {}
for dept in all_depts:
    dw = dept_debt_weighted[dept]['weighted_days']
    da = dept_debt_weighted[dept]['total_debt']
    rw = dept_rec_weighted[dept]['weighted_days']
    ra = dept_rec_weighted[dept]['total_rec']
    
    total_weighted = dw + rw
    total_amt = da + ra
    
    if total_amt > 0:
        cycle = total_weighted / total_amt
    else:
        cycle = 0
    
    results[dept] = round(cycle, 1)
    total_debt_weighted += dw
    total_rec_weighted += rw
    total_debt_amt += da
    total_rec_amt += ra
    
    print(f"  {dept}: 欠款加权={dw/10000:.0f}万元天, 回款加权={rw/10000:.0f}万元天, "
          f"分母={total_amt/10000:.2f}万, 周期={cycle:.1f}天")

# 全大区汇总
total_weighted_all = total_debt_weighted + total_rec_weighted
total_amt_all = total_debt_amt + total_rec_amt
if total_amt_all > 0:
    total_cycle = total_weighted_all / total_amt_all
else:
    total_cycle = 0

print(f"\n=== 全大区汇总 ===")
print(f"  欠款总额: {total_debt_amt/10000:.2f}万")
print(f"  认款总额: {total_rec_amt/10000:.2f}万")
print(f"  欠款加权天数合计: {total_debt_weighted/10000:.0f}万元天")
print(f"  回款加权天数合计: {total_rec_weighted/10000:.0f}万元天")
print(f"  整体回款周期: {total_cycle:.1f}天")
print()
print("dept_cycle_results:", results)
