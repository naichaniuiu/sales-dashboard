# -*- coding: utf-8 -*-
"""
从 26财年Q1业绩 Excel 中提取 部门→销售员→客户 的业绩下钻数据，
生成 deptCustPerfData JSON 并注入到 Q1 看板 HTML 中。
"""
import openpyxl, json
from collections import defaultdict

EXCEL = r'C:\Users\wm881\Downloads\业绩 欠款看板.xlsx'
HTML  = r'C:\Users\wm881\WorkBuddy\20260513090923\中西部大区26财年Q1数据看板.html'

# ===== 1. 读取 Excel =====
wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb['26财年Q1业绩']

# 三级部门=31, 销售员=27, 客户名称=14, 业绩总金额=20, 回款金额=21, 欠款金额=22, 协同比例=43
# 构建 dept -> sales -> { customer: { perf, collect, debt, orders } }
dept_map = {}  # dept -> sales -> cust -> {perf, collect, debt, orders}

for r in range(2, ws.max_row + 1):
    dept   = ws.cell(r, 31).value  # 三级部门
    sales  = ws.cell(r, 27).value  # 销售员名称
    cust   = ws.cell(r, 14).value  # 客户名称
    perf   = ws.cell(r, 20).value or 0  # 业绩总金额
    coll   = ws.cell(r, 21).value or 0  # 回款金额
    debt   = ws.cell(r, 22).value or 0  # 欠款金额
    ratio  = ws.cell(r, 43).value or 1  # 协同比例

    if not dept or not sales or not cust:
        continue
    if not isinstance(perf, (int, float)):
        perf = 0
    if not isinstance(coll, (int, float)):
        coll = 0
    if not isinstance(debt, (int, float)):
        debt = 0
    if not isinstance(ratio, (int, float)):
        ratio = 1

    dept_map.setdefault(dept, {}).setdefault(sales, {}).setdefault(cust, {'perf': 0.0, 'collect': 0.0, 'debt': 0.0, 'orders': 0})
    entry = dept_map[dept][sales][cust]
    entry['perf']   += perf * ratio
    entry['collect'] += coll * ratio
    entry['debt']   += debt * ratio
    entry['orders'] += 1

# ===== 2. 转换为 deptCustPerfData 结构 =====
deptCustPerfData = {}
for dept, sales_map in dept_map.items():
    sales_list = []
    for sales, cust_map in sales_map.items():
        custs = []
        total_perf = 0
        total_coll = 0
        total_debt = 0
        for cust, vals in cust_map.items():
            p = round(vals['perf'] / 10000, 2)
            c = round(vals['collect'] / 10000, 2)
            d = round(vals['debt'] / 10000, 2)
            total_perf += p
            total_coll += c
            total_debt += d
            custs.append({
                'name': cust,
                'perf': p,
                'collect': c,
                'debt': d,
                'orders': vals['orders']
            })
        # Sort customers by perf desc
        custs.sort(key=lambda x: x['perf'], reverse=True)
        sales_list.append({
            'name': sales,
            'total': round(total_perf, 2),
            'collect': round(total_coll, 2),
            'debt': round(total_debt, 2),
            'cust_count': len(custs),
            'custs': custs
        })
    # Sort salespeople by perf desc
    sales_list.sort(key=lambda x: x['total'], reverse=True)
    deptCustPerfData[dept] = sales_list

# Print summary
total_perf = 0
total_custs = 0
for dept, sl in deptCustPerfData.items():
    dept_perf = sum(s['total'] for s in sl)
    dept_custs = sum(s['cust_count'] for s in sl)
    total_perf += dept_perf
    total_custs += dept_custs
    print(f"{dept}: {len(sl)} sales, {dept_custs} custs, perf={dept_perf:.2f}万")
print(f"\n总计: 业绩={total_perf:.2f}万, {total_custs}个客户")

# ===== 3. 注入到 HTML =====
json_str = json.dumps(deptCustPerfData, ensure_ascii=False, separators=(',', ':'))
print(f"\nJSON size: {len(json_str)} chars")

with open(HTML, 'r', encoding='utf-8') as f:
    html = f.read()

# Check if already injected
if 'deptCustPerfData' in html:
    # Replace existing
    import re
    html = re.sub(
        r'let deptCustPerfData\s*=\s*\{[^;]*\};',
        'let deptCustPerfData = ' + json_str + ';',
        html, count=1
    )
    print("Replaced existing deptCustPerfData")
else:
    # Insert after deptCustData
    marker = 'let deptCustData = '
    idx = html.find(marker)
    if idx == -1:
        print("ERROR: Cannot find deptCustData marker!")
        exit(1)
    # Find end of the deptCustData assignment (semicolon)
    semi_idx = html.index(';', idx)
    insert_pos = semi_idx + 1
    html = html[:insert_pos] + '\nlet deptCustPerfData = ' + json_str + ';' + html[insert_pos:]
    print("Inserted deptCustPerfData after deptCustData")

with open(HTML, 'w', encoding='utf-8') as f:
    f.write(html)
print(f"\nInjected deptCustPerfData ({len(json_str)} chars) into HTML")
